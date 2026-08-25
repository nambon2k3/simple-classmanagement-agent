"""Read a student roster out of an uploaded Excel or CSV file.

Teachers keep rosters in whatever spreadsheet they already have, so the parser
recognises a header row by keyword (English or Vietnamese) and falls back to
column order when the sheet has no header at all.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field

#: Header keywords per target field, checked as substrings of the lower-cased cell.
#: Order matters: "Student name" must map to the name column, not the ID column.
_HEADER_KEYWORDS: dict[str, tuple[str, ...]] = {
    "email": ("email", "mail"),
    "phone": ("phone", "sdt", "so dien thoai", "số điện thoại", "dien thoai", "điện thoại"),
    "note": ("note", "ghi chu", "ghi chú", "comment"),
    "full_name": ("name", "ten", "tên", "hoc sinh", "học sinh"),
    "student_code": ("code", "id", "mssv", "ma so", "mã số", "ma hs"),
}

#: Column order used when the sheet has no recognisable header row.
_POSITIONAL = ("student_code", "full_name", "email", "phone", "note")

_MAX_ROWS = 500


@dataclass(frozen=True)
class RosterRow:
    """One student parsed out of the uploaded file."""

    row_number: int
    student_code: str
    full_name: str
    email: str | None = None
    phone: str | None = None
    note: str | None = None


@dataclass
class RosterFile:
    """Everything the importer needs from one uploaded file."""

    rows: list[RosterRow] = field(default_factory=list)
    problems: list[str] = field(default_factory=list)


class RosterFileError(ValueError):
    """The uploaded file could not be read at all."""


def parse_roster_file(filename: str, data: bytes) -> RosterFile:
    """Parse an ``.xlsx``/``.xlsm`` or ``.csv`` roster into rows.

    Args:
        filename: Original name, used only to pick the reader.
        data: Raw file bytes.

    Returns:
        Parsed rows plus a human-readable problem per unusable line.

    Raises:
        RosterFileError: If the extension is unsupported or the file is corrupt.
    """
    lowered = filename.lower()
    if lowered.endswith((".xlsx", ".xlsm")):
        table = _read_excel(data)
    elif lowered.endswith(".csv"):
        table = _read_csv(data)
    else:
        raise RosterFileError("Please upload an .xlsx or .csv file.")
    return _to_rows(table)


def _read_excel(data: bytes) -> list[list[str]]:
    """Return the first worksheet as a table of trimmed strings."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise RosterFileError("Excel support is not installed on the server.") from exc

    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise RosterFileError("That file is not a readable Excel workbook.") from exc
    try:
        sheet = workbook.worksheets[0]
        return [[_cell_text(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _read_csv(data: bytes) -> list[list[str]]:
    """Return a CSV file as a table of trimmed strings."""
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("latin-1")
    return [[_cell_text(cell) for cell in row] for row in csv.reader(io.StringIO(text))]


def _cell_text(value: object) -> str:
    """Render a spreadsheet cell as text, keeping phone numbers intact."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _to_rows(table: list[list[str]]) -> RosterFile:
    """Map a raw table onto roster rows, skipping blanks and reporting gaps."""
    populated = [(index, row) for index, row in enumerate(table, start=1) if any(row)]
    if not populated:
        raise RosterFileError("That file has no rows.")

    mapping = _header_mapping(populated[0][1])
    if mapping is None:
        # No header to go on, so fall back to column order — which is only
        # usable when there is an ID column and a name column to read.
        if max(len(row) for _, row in populated) < 2:
            raise RosterFileError(
                "I need a student ID column and a name column. "
                "Name them 'Student ID' and 'Full name'."
            )
        mapping = dict(enumerate(_POSITIONAL))
        body = populated
    else:
        body = populated[1:]

    parsed = RosterFile()
    for row_number, row in body[:_MAX_ROWS]:
        values = {
            field_name: row[column].strip() if column < len(row) else ""
            for column, field_name in mapping.items()
        }
        code = values.get("student_code", "")
        name = values.get("full_name", "")
        if not code or not name:
            missing = "student ID" if not code else "name"
            parsed.problems.append(f"Row {row_number}: missing {missing}.")
            continue
        parsed.rows.append(
            RosterRow(
                row_number=row_number,
                student_code=code,
                full_name=name,
                email=values.get("email") or None,
                phone=values.get("phone") or None,
                note=values.get("note") or None,
            )
        )

    if len(body) > _MAX_ROWS:
        parsed.problems.append(
            f"Only the first {_MAX_ROWS} rows were read; the file has {len(body)}."
        )
    return parsed


def _header_mapping(row: list[str]) -> dict[int, str] | None:
    """Map column index to field name, or ``None`` when this is not a header."""
    mapping: dict[int, str] = {}
    for column, cell in enumerate(row):
        text = cell.strip().lower()
        if not text:
            continue
        for field_name, keywords in _HEADER_KEYWORDS.items():
            if field_name in mapping.values():
                continue
            if any(keyword in text for keyword in keywords):
                mapping[column] = field_name
                break
    if "student_code" in mapping.values() and "full_name" in mapping.values():
        return mapping
    return None
